from collections import defaultdict
import time
import serial
from os import makedirs
import tkinter as tk
from tkinter import font as tkFont
import serial.tools.list_ports
from openpyxl import Workbook, load_workbook
from tkinter import Toplevel
import pygetwindow as gw
import pyautogui
from PIL import ImageGrab
import time
import subprocess

# -------------------------------------------------------------------------------------------
DEBUG = True
# DEBUG = False

# ===========================================================================================
class PLCReader:
# -------------------------------------------------------------------------------------------
    def __init__(self):
        self.stationNo = b'01' #국번(gukbun)
        self.ser = None
        self.isConnected = False
        self.PLCAddr = {
            'isRunning':  'P020',
            'PrgNo':      'D131',
            'BondName':   'D132',
            'LotNo':      'D140',
            'RPM':        'D120',
            'TimeSet':    'D122',
            'TimeAct':    'D220',
            }
        self.PLCDataName= list(self.PLCAddr.keys())      
        self.PLCAddrSize = defaultdict(lambda:1, { # word length: DEFAULT 1W
            'BondName':   8, # BondName: 8W
            'LotNo':      5, # LotNo:5W
            })
    # -------------------------------------------------------------------------------------------
    def connect(self):
        try:
            if not self.ser.is_open: raise
        except:
            ports = serial.tools.list_ports.comports()
            if len(ports) == 1:
                portName = ports[0].device
            elif len(ports) == 0:
                print("연결된 장치 없음")
            else:
                for portInfo in ports:
                    print(f" - Device Name: {portInfo.device}")
                    print(f"   Serial Number: {portInfo.serial_number}")
                    print(f"   Vendor ID: {portInfo.vid}")
                    print(f"   Product ID: {portInfo.pid}")
                    print(f"   Description: {portInfo.description}")
                    print("\n")
                else:
                    portName = input("Device Name 입력: ").strip()
            try:
                self.ser = serial.Serial(portName,115200,timeout=1)
                self.isConnected = True
            except:
                self.ser = None
                self.isConnected = False
        else:
            self.isConnected = True
        finally:
            return self.isConnected
    # --------------------------
    def disconnect(self):
        try:
            self.ser.close()
        except:...
        finally:
            self.ser = None
            self.isConnected = False
            return self.isConnected
    # -------------------------------------------------------------------------------------------    
    def getIsRunning(self) -> bool:
        res = self.getPLCRecv('isRunning')
        print('isRunning' + ': ' + res)
        if res == '0': return False
        else:          return True
    # --------------------------
    def getResList(self) -> list:
        resList = []
        for dataName in self.PLCDataName:
            res = self.getPLCRecv(dataName)
            if 'Time' in dataName:
                res = format(int(res)/10, '.1f')
            resList.append(res)
            print(dataName + ': ' + res)
        if resList[0] == '1':
            resList[0] = 'RUN'
        elif resList[0] == '0':
            resList[0] = 'STOP'
        else: ...
        return resList
    # --------------------------
    def getTimeAct(self) -> str:
        res = self.getPLCRecv('TimeAct')
        res = format(int(res)/10, '.1f')
        print('TimeAct' + ': ' + res)
        return res
    # -------------------------------------------------------------------------------------------
    def MakeReadCmd(self, dataName:str, isBit:bool=False) -> str:
        readAddrStr = self.PLCAddr[dataName]
        if isBit:   dataType = b'X'
        else:       dataType = b'W'
        readAddrByte = b'%' + readAddrStr[0].encode('ascii') + dataType + readAddrStr[1:].encode('ascii')
        # --------------------------
        wordLength = self.PLCAddrSize[dataName]
        wordLengthByte = (hex(wordLength)).replace('0x','').replace('0X','').zfill(2).encode('ascii')
        # --------------------------
        if wordLength == 1: readCmd = b'RSS0106' + readAddrByte # RSS
        else: readCmd = b'RSB06' + readAddrByte + wordLengthByte # RSB
        return b'\x05' + self.stationNo + readCmd + b'\x04'
    # -------------------------------------------------------------------------------------------
    def getPLCRecv(self, dataName:str=None):
        if DEBUG:
            if dataName=='isRunning': return 'RUN'            
            if int(time.mktime(time.localtime())) % 2:
                recv = b'\x0601RSB010233343536000000000000\x03'
            else:
                recv = b'\x0601RSS0102000A\x03'
        else:
            if dataName=='isRunning':
                cmd = self.MakeReadCmd(dataName,True)
            else:
                cmd = self.MakeReadCmd(dataName)
            # --------------------------
            firstErrorTime, errorTime = None, 0
            while True:
                try:
                    self.ser.write(cmd)
                    recv = self.ser.readline() # ex) recv = b'\x0601RSS0102000A\x03'    
                except Exception as e:
                    if not firstErrorTime:
                        firstErrorTime = time.time()
                    else:
                        errorTime = time.time() - firstErrorTime
                    # --------------------------
                    if errorTime > 300: # wait 5 min
                        ...
                    else: time.sleep(1)
                else:
                    break
        # --------------------------
        def decodeResult() -> str:
            recvStr = recv.decode('ascii')
            resultStr = ''        
            if recv[1:3] != self.stationNo: ...
            elif recvStr.startswith('\x06'): # ACK
                if recvStr[3:6] == 'RSS':
                    resultStr = str(int(recvStr[10:-1],16))
                elif recvStr[3:6] == 'RSB':
                    hexStr = recvStr[10:-1]
                    resultStr = ''.join([chr(int(hexStr[i+2:i+4], 16))+chr(int(hexStr[i:i+2], 16)) for i in range(0, len(hexStr), 4)])
                else: ... # not read result
            else: ... # not ACK
            return resultStr.replace('\x00','').strip()
        return decodeResult()

# ===========================================================================================
class MonitorWindow:
# -------------------------------------------------------------------------------------------
    def __init__(self,reader:PLCReader):
        self.reader = reader
        self.window = tk.Tk()
        self.window.title("MixerMonitor")
        self.window.geometry("700x400")
        self.font = tkFont.Font(family='Helvetica', size=20)
        self.entryList, self.titleLabel = self.setWindow()
        self.running,self.wasRunning = False, False
        self.resList = []
        # --------------------------
        self.connect()
        self.window.mainloop()
    # -------------------------------------------------------------------------------------------
    def setInputWindow(self):
        inputwindow = Toplevel(self.window)
        inputwindow.title("데이터 직접 입력")
        # --------------------------
        entry = tk.Entry(inputwindow, width=40, font=self.font)
        entry.pack(pady=1,)
        entry.bind("<Return>", lambda event: self.loadEntryValue(entry))
        # --------------------------
        tk.Button(inputwindow, text="출력하기", command=lambda: self.loadEntryValue(entry)).pack(side=tk.LEFT, padx=5)
        tk.Button(inputwindow, text="내용삭제", command=lambda: self.clearEntry(entry)).pack(side=tk.LEFT, padx=5)
    # --------------------------
    def loadEntryValue(self, entry):
        value = entry.get()
        print("입력값:", value)
        self.resList = value.split('\t')
        self.updateEntries(self.resList)

    def clearEntry(self, entry):
        entry.delete(0, tk.END)
    # -------------------------------------------------------------------------------------------
    def updateTitleLabel(self):
        try:    portName = self.reader.ser.name
        except: portName = 'disconn.'
        finally:
            self.titleLabel['text'] = portName
    # --------------------------
    def disconnect(self):
        self.reader.disconnect()
        self.updateTitleLabel()

    def connect(self):
        self.reader.connect()
        self.updateTitleLabel()
        if self.reader.ser: 
            self.update()
    # -------------------------------------------------------------------------------------------
    def printRequest(self, imagePath):
        print(imagePath)
        subprocess.run(["start", "mspaint", "/p", imagePath.split('/')[-1]], shell=True, cwd='./capture')

    def captureWindow(self):
        targetWindow = gw.getWindowsWithTitle(self.window.title())[0]
        x, y, w, h = targetWindow.left, targetWindow.top, targetWindow.width, targetWindow.height

        # pyautogui.moveTo(x, y)
        time.sleep(0.5)
        screenshot = ImageGrab.grab(bbox=(x+10, y+3, x + w -10, y + h -10))

        if len(self.resList) < len(self.entryList):
            nowStr = time.strftime("%Y%m%d%H%M%S", time.localtime())
        else:
            nowStr = self.resList[0].replace('-','').replace(' ','').replace(':','')
        captureName = f'{nowStr}.png'
        capturePath = './capture/' + captureName
        # --------------------------
        makedirs('./capture', exist_ok=True)
        screenshot.save(capturePath)
        return captureName

    def captureAndPrint(self):
        captureName = self.captureWindow()
        self.printRequest(captureName)

    # -------------------------------------------------------------------------------------------
    def setWindow(self):
        entryList = []
        # --------------------------
        menubar = tk.Menu(self.window)
        self.window.config(menu=menubar)

        def newMenu(menuName:str,cmdDict:dict):
            menu = tk.Menu(menubar, tearoff=0)
            for cmd in cmdDict:
                menu.add_command(label=cmd, command=cmdDict[cmd])
            menubar.add_cascade(label=menuName, menu=menu)
            return menu

        newMenu("캡쳐",{ "이미지로 저장" : self.captureWindow, "바로 인쇄" : self.captureAndPrint})
        newMenu("뷰어",{ "데이터 직접 입력": self.setInputWindow})
        newMenu("연결",{ "연결하기" : self.connect, "연결해제" : self.disconnect})

        # --------------------------
        def newLabel(text:str,x:int,y:int,xSize:int=1,ySize:int=1):
            label = tk.Label(self.window, text=text, font=self.font)
            label.grid(row=y, rowspan=ySize, column=x, columnspan=xSize, padx=5, pady=5)
            return label

        def newEntry(x:int,y:int,xSize:int=1,ySize:int=1,sort:str='center',style:str='default'):
            if style in ['short']:
                width, padx, pady = 5,15,15
            elif style in ['long']:
                width, padx, pady = 21,5,5
            else:
                width, padx, pady = 10,5,5
            entry = tk.Entry(self.window, width=width, font=self.font, justify=sort)
            entry.grid(row=y, rowspan=ySize, column=x, columnspan=xSize, padx=padx, pady=pady)
            entryList.append(entry)
            return entry
        # --------------------------
        if self.reader.ser: titleText = self.reader.ser.port
        else:               titleText = 'disconnect'
        titleLabel = newLabel(titleText,0,0)
        titleLabel["fg"] = "blue"

        newEntry(1,0,2,1,'left','long'); newEntry(2,3,2,2,style='short')   # nowTime, isRunning
        newLabel('PRG. NO',0,1);     newEntry(0,2,style='short')
        newLabel('BOND NAME',1,1,2); newEntry(1,2,2,1,'left','long')
        newLabel('LOT NO',0,3);      newEntry(1,3)
        newLabel('RPM',0,4);         newEntry(1,4)
        newLabel('Time(Min)',0,5);   newEntry(1,5); newEntry(2,5)
        return entryList, titleLabel

    # -------------------------------------------------------------------------------------------
    def updateEntries(self, resList):
        if (resList == []) or (resList == ['']):
            for entry in self.entryList:
                entry.delete(0, tk.END)
        else:
            if len(resList) < len(self.entryList):
                nowStr = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                resList = [nowStr]+resList
            for entry, res in zip(self.entryList, resList):
                entry.delete(0, tk.END)
                entry.insert(0, res)
        self.window.update()

    def update(self):
        if self.reader.isConnected: 
            self.running = self.reader.getIsRunning()
            if self.running != self.wasRunning:             # start or stop
                if self.running:
                    self.titleLabel['text'] = "Loading..."
                    self.updateEntries([])
                    # --------------------------
                    self.resList = self.reader.getResList()
                    self.titleLabel['text'] = self.reader.ser.port
                else:
                    self.resList[0] = 'STOP'
                    self.resList[-1] = self.reader.getTimeAct()
                self.updateEntries(self.resList)
                appendListToExcel(self.resList)
                self.wasRunning = self.running # update
            elif self.running:                              # running
                self.resList[-1] = self.reader.getTimeAct()
                self.updateEntries(self.resList)
            else: ...                                       # stopped
        self.window.after(500, self.update)
    # -------------------------------------------------------------------------------------------

# ===========================================================================================
def appendListToExcel(dataList):
    fileHeader = ['시간','RUN/STOP','PRG. NO','BOND NAME','LOT NO','RPM','설정시간','동작시간']
    resFileName = time.strftime("./output/%Y_%m_%d.xlsx", time.localtime())
    nowStr = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) 
    dataList = [nowStr]+dataList 
    try:
        try:
            workbook = load_workbook(resFileName)
        except FileNotFoundError:
            makedirs('./output', exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(fileHeader)
        else:
            sheet = workbook.active
        finally:
            sheet.append(dataList)
            workbook.save(resFileName)
            print("데이터가 성공적으로 추가되었습니다.")
    except Exception as e:
        print(f"오류 발생: {str(e)}")

# ===========================================================================================
if __name__ == "__main__":
    reader = PLCReader()
    mw = MonitorWindow(reader)
# ===========================================================================================
